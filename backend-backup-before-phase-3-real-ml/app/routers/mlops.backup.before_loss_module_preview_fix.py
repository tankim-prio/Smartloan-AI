from datetime import datetime
from pathlib import Path
from typing import Any
import json
import csv
import pickle
import shutil
import uuid
import itertools

from fastapi import APIRouter, Body, Depends, File, Form, HTTPException, UploadFile
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.core.security import get_current_user
from app.database import get_db
from app.models.application import Application


router = APIRouter(prefix="/mlops", tags=["MLOps"])


UPLOAD_ROOT = Path("uploads") / "mlops"
MAX_UPLOAD_BYTES = 300 * 1024 * 1024


def now_text() -> str:
    return datetime.utcnow().isoformat(timespec="seconds")


def is_admin(current_user) -> bool:
    return str(getattr(current_user, "role", "") or "").lower() == "admin"


def check_admin(current_user) -> None:
    if not is_admin(current_user):
        raise HTTPException(status_code=403, detail="Admin access required")



def ensure_mlops_tables(db: Session) -> None:
    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS ml_models (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                model_name TEXT,
                version TEXT DEFAULT 'v1',
                model_type TEXT DEFAULT 'uploaded_model',
                description TEXT,
                status TEXT DEFAULT 'registered',
                is_active INTEGER DEFAULT 0,
                dataset_path TEXT,
                dataset_original_name TEXT,
                model_path TEXT,
                model_original_name TEXT,
                accuracy REAL DEFAULT 0,
                f1_score REAL DEFAULT 0,
                created_at TEXT,
                updated_at TEXT,
                deployed_at TEXT,
                activated_at TEXT
            )
            """
        )
    )

    db.execute(
        text(
            """
            CREATE TABLE IF NOT EXISTS ml_predictions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                application_id INTEGER,
                model_id INTEGER,
                result TEXT,
                risk_level TEXT,
                confidence REAL,
                reason TEXT,
                features_json TEXT,
                created_at TEXT
            )
            """
        )
    )

    db.commit()

    def add_missing_columns(table_name: str, columns: dict[str, str]) -> None:
        existing_rows = db.execute(text(f"PRAGMA table_info({table_name})")).mappings().all()
        existing_columns = {row["name"] for row in existing_rows}

        for column_name, column_type in columns.items():
            if column_name not in existing_columns:
                db.execute(text(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {column_type}"))

        db.commit()

    add_missing_columns(
        "ml_models",
        {
            "model_name": "TEXT",
            "version": "TEXT DEFAULT 'v1'",
            "model_type": "TEXT DEFAULT 'uploaded_model'",
            "description": "TEXT",
            "status": "TEXT DEFAULT 'registered'",
            "is_active": "INTEGER DEFAULT 0",
            "dataset_path": "TEXT",
            "dataset_original_name": "TEXT",
            "model_path": "TEXT",
            "model_original_name": "TEXT",
            "accuracy": "REAL DEFAULT 0",
            "f1_score": "REAL DEFAULT 0",
            "created_at": "TEXT",
            "updated_at": "TEXT",
            "deployed_at": "TEXT",
            "activated_at": "TEXT",
        },
    )

    add_missing_columns(
        "ml_predictions",
        {
            "application_id": "INTEGER",
            "model_id": "INTEGER",
            "result": "TEXT",
            "risk_level": "TEXT",
            "confidence": "REAL",
            "reason": "TEXT",
            "features_json": "TEXT",
            "created_at": "TEXT",
        },
    )


def safe_text(value: Any, default: str = "") -> str:
    text_value = str(value or "").strip()
    return text_value if text_value else default


def safe_number(value: Any, default: float = 0) -> float:
    try:
        if value in [None, ""]:
            return default
        return float(str(value).replace(",", "").strip())
    except Exception:
        return default


def get_application_or_404(db: Session, application_id: int) -> Application:
    application = db.query(Application).filter(Application.id == application_id).first()

    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    return application


def application_to_features(application: Application) -> dict:
    age = safe_number(getattr(application, "age", None), 0)
    monthly_income = safe_number(getattr(application, "monthly_income", None), 0)
    occupation = safe_text(getattr(application, "occupation", ""), "")
    phone = safe_text(getattr(application, "phone", ""), "")
    email = safe_text(getattr(application, "email", ""), "")
    address = safe_text(getattr(application, "address", ""), "")

    return {
        "age": age,
        "monthly_income": monthly_income,
        "occupation": occupation,
        "phone": phone,
        "email": email,
        "address": address,
        "has_occupation": 1 if occupation else 0,
        "has_phone": 1 if phone else 0,
        "has_email": 1 if email else 0,
        "has_address": 1 if address else 0,
    }


def baseline_prediction(features: dict) -> dict:
    age = safe_number(features.get("age"), 0)
    income = safe_number(features.get("monthly_income"), 0)
    has_occupation = int(features.get("has_occupation") or 0)
    has_phone = int(features.get("has_phone") or 0)
    has_email = int(features.get("has_email") or 0)

    score = 0

    if income >= 50000:
        score += 40
    elif income >= 25000:
        score += 25
    elif income > 0:
        score += 10

    if 22 <= age <= 60:
        score += 20
    elif age > 0:
        score += 8

    if has_occupation:
        score += 20

    if has_phone:
        score += 10

    if has_email:
        score += 10

    confidence = max(55, min(95, score))

    if score >= 70:
        return {
            "result": "recommended_approval",
            "risk_level": "low",
            "confidence": confidence,
            "reason": "Monthly income is strong, applicant age is within stable working range, and required contact/occupation information is complete.",
        }

    if score >= 45:
        return {
            "result": "manual_review",
            "risk_level": "medium",
            "confidence": confidence,
            "reason": "Applicant information is partially strong, but income or profile completeness needs admin review.",
        }

    return {
        "result": "not_recommended",
        "risk_level": "high",
        "confidence": confidence,
        "reason": "Applicant profile has weak income or incomplete required information, so manual rejection/review is recommended.",
    }




def get_predictable_model_object(model_object):
    if model_object is None:
        return None

    try:
        if callable(getattr(model_object, "predict", None)):
            return model_object
    except Exception:
        pass

    if isinstance(model_object, dict):
        for key in ["model", "estimator", "classifier", "pipeline", "trained_model"]:
            possible_model = model_object.get(key)

            try:
                if possible_model is not None and callable(getattr(possible_model, "predict", None)):
                    return possible_model
            except Exception:
                pass

    return None


def build_model_feature_value(feature_name: str, features: dict) -> float:
    name = str(feature_name or "").strip().lower()

    age = safe_number(features.get("age"), 0)
    monthly_income = safe_number(features.get("monthly_income"), 0)

    # Direct Apply page fields
    if name in ["age", "applicant_age"]:
        return age

    if name in ["monthly_income", "income", "salary", "monthly_salary"]:
        return monthly_income

    # Missing fields that the uploaded sklearn model may require.
    # These defaults keep prediction stable when Apply page does not collect them yet.
    if name in ["loan_amount", "requested_loan_amount", "amount"]:
        # Safe estimate: if income exists, use 6 months of income; otherwise default 100000.
        return monthly_income * 6 if monthly_income > 0 else 100000

    if name in ["credit_score", "cibil_score", "score"]:
        # Neutral-good default score.
        return safe_number(features.get("credit_score"), 650)

    if name in ["employment_years", "years_employed", "work_experience", "experience_years"]:
        # If occupation exists, assume at least 2 years; otherwise 0.
        return safe_number(features.get("employment_years"), 2 if safe_text(features.get("occupation")) else 0)

    if name in ["existing_debt", "current_debt", "debt"]:
        return safe_number(features.get("existing_debt"), 0)

    # Existing engineered fields from Apply page
    if name in ["has_occupation"]:
        return 1 if safe_text(features.get("occupation")) else 0

    if name in ["has_phone"]:
        return 1 if safe_text(features.get("phone")) else 0

    if name in ["has_email"]:
        return 1 if safe_text(features.get("email")) else 0

    if name in ["has_address"]:
        return 1 if safe_text(features.get("address")) else 0

    # Unknown feature: use 0 so model call does not crash.
    return safe_number(features.get(name), 0)


def build_model_input(model_object, features: dict):
    feature_names = []

    try:
        raw_feature_names = getattr(model_object, "feature_names_in_", None)

        if raw_feature_names is not None:
            feature_names = [str(item) for item in list(raw_feature_names)]
    except Exception:
        feature_names = []

    if not feature_names:
        feature_names = [
            "age",
            "monthly_income",
            "has_occupation",
            "has_phone",
            "has_email",
            "has_address",
        ]

    row = {
        feature_name: build_model_feature_value(feature_name, features)
        for feature_name in feature_names
    }

    try:
        import pandas as pd

        return pd.DataFrame([row], columns=feature_names), feature_names, row
    except Exception:
        return [[row[feature_name] for feature_name in feature_names]], feature_names, row


def normalize_uploaded_prediction(raw_prediction, probability=None) -> dict:
    prediction_text = str(raw_prediction).strip().lower()

    if prediction_text in ["1", "true", "approved", "approve", "recommended_approval", "low", "yes"]:
        return {
            "result": "recommended_approval",
            "risk_level": "low",
            "confidence": probability or 85,
        }

    if prediction_text in ["2", "manual_review", "review", "medium"]:
        return {
            "result": "manual_review",
            "risk_level": "medium",
            "confidence": probability or 75,
        }

    return {
        "result": "not_recommended",
        "risk_level": "high",
        "confidence": probability or 70,
    }


def try_uploaded_model_prediction(model_path: str | None, features: dict) -> dict | None:
    if not model_path:
        return None

    path = Path(model_path)

    if not path.exists() or not path.is_file():
        return None

    try:
        model_object, loader, load_errors = load_model_object_safely(path)
        predictable_model = get_predictable_model_object(model_object)

        if predictable_model is None:
            return None

        model_input, feature_names, feature_row = build_model_input(predictable_model, features)

        raw_prediction = predictable_model.predict(model_input)[0]

        probability = None

        predict_proba = getattr(predictable_model, "predict_proba", None)

        if callable(predict_proba):
            try:
                probabilities = list(predict_proba(model_input)[0])
                probability = float(max(probabilities)) * 100
            except Exception:
                probability = None

        normalized = normalize_uploaded_prediction(raw_prediction, probability)

        normalized["reason"] = (
            f"Prediction generated by active uploaded model using {loader}. "
            f"Mapped features: {', '.join(feature_names)}."
        )

        normalized["mapped_features"] = feature_row

        return normalized

    except Exception:
        return None

def save_upload_file(upload_file: UploadFile, folder: Path) -> tuple[str, str]:
    folder.mkdir(parents=True, exist_ok=True)

    original_name = upload_file.filename or "uploaded_file"
    suffix = Path(original_name).suffix or ".bin"
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    file_path = folder / stored_name

    total_size = 0
    chunk_size = 1024 * 1024

    try:
        upload_file.file.seek(0)
    except Exception:
        pass

    with open(file_path, "wb") as buffer:
        while True:
            chunk = upload_file.file.read(chunk_size)

            if not chunk:
                break

            total_size += len(chunk)

            if total_size > MAX_UPLOAD_BYTES:
                try:
                    file_path.unlink(missing_ok=True)
                except Exception:
                    pass

                raise HTTPException(
                    status_code=413,
                    detail="File is too large. Maximum supported upload size is 300MB.",
                )

            buffer.write(chunk)

    return str(file_path), original_name


def serialize_model(row: Any) -> dict:
    return {
        "id": row["id"],
        "model_name": row["model_name"],
        "version": row["version"],
        "model_type": row["model_type"],
        "description": row["description"],
        "status": row["status"],
        "is_active": bool(row["is_active"]),
        "dataset_path": row["dataset_path"],
        "dataset_original_name": row["dataset_original_name"],
        "model_path": row["model_path"],
        "model_original_name": row["model_original_name"],
        "accuracy": row["accuracy"],
        "f1_score": row["f1_score"],
        "created_at": row["created_at"],
        "updated_at": row["updated_at"],
        "deployed_at": row["deployed_at"],
        "activated_at": row["activated_at"],
    }


def get_model_or_404(db: Session, model_id: int) -> dict:
    ensure_mlops_tables(db)

    row = db.execute(
        text("SELECT * FROM ml_models WHERE id = :id"),
        {"id": model_id},
    ).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="ML model not found")

    return dict(row)


def get_active_model(db: Session) -> dict | None:
    ensure_mlops_tables(db)

    row = db.execute(
        text("SELECT * FROM ml_models WHERE is_active = 1 ORDER BY id DESC LIMIT 1")
    ).mappings().first()

    return dict(row) if row else None


@router.get("/health")
def mlops_health(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    return {
        "ok": True,
        "message": "MLOps backend is ready.",
    }



@router.post("/models/upload")
def upload_ml_model_package(
    model_name: str = Form(...),
    version: str = Form(default="v1"),
    model_type: str = Form(default="uploaded_model"),
    description: str = Form(default=""),
    dataset_file: UploadFile | None = File(default=None),
    model_file: UploadFile | None = File(default=None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    created_at = now_text()

    result = db.execute(
        text(
            """
            INSERT INTO ml_models (
                model_name,
                version,
                model_type,
                description,
                status,
                is_active,
                accuracy,
                f1_score,
                created_at,
                updated_at
            )
            VALUES (
                :model_name,
                :version,
                :model_type,
                :description,
                'registered',
                0,
                0,
                0,
                :created_at,
                :updated_at
            )
            """
        ),
        {
            "model_name": model_name,
            "version": version,
            "model_type": model_type,
            "description": description,
            "created_at": created_at,
            "updated_at": created_at,
        },
    )

    db.commit()

    model_id = result.lastrowid
    model_folder = UPLOAD_ROOT / "models" / str(model_id)

    dataset_path = None
    dataset_original_name = None
    model_path = None
    model_original_name = None

    if dataset_file is not None:
        dataset_path, dataset_original_name = save_upload_file(dataset_file, model_folder / "dataset")

    if model_file is not None:
        model_path, model_original_name = save_upload_file(model_file, model_folder / "model")

    db.execute(
        text(
            """
            UPDATE ml_models
            SET dataset_path = :dataset_path,
                dataset_original_name = :dataset_original_name,
                model_path = :model_path,
                model_original_name = :model_original_name,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "dataset_path": dataset_path,
            "dataset_original_name": dataset_original_name,
            "model_path": model_path,
            "model_original_name": model_original_name,
            "updated_at": now_text(),
            "id": model_id,
        },
    )

    db.commit()

    model = get_model_or_404(db, model_id)

    return {
        "created": True,
        "message": "Model package uploaded successfully.",
        "model": serialize_model(model),
    }


@router.get("/models")
def list_ml_models(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    rows = db.execute(
        text("SELECT * FROM ml_models ORDER BY id DESC")
    ).mappings().all()

    return [serialize_model(dict(row)) for row in rows]


@router.get("/models/active")
def get_active_ml_model(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    model = get_active_model(db)

    if not model:
        return {
            "active": False,
            "model": None,
        }

    return {
        "active": True,
        "model": serialize_model(model),
    }


@router.patch("/models/{model_id}/deploy")
def deploy_ml_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    get_model_or_404(db, model_id)

    db.execute(
        text(
            """
            UPDATE ml_models
            SET status = 'deployed',
                deployed_at = :deployed_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "deployed_at": now_text(),
            "updated_at": now_text(),
            "id": model_id,
        },
    )

    db.commit()

    return {
        "deployed": True,
        "model": serialize_model(get_model_or_404(db, model_id)),
    }


@router.patch("/models/{model_id}/set-active")
def set_active_ml_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    model = get_model_or_404(db, model_id)

    if model["status"] != "deployed":
        raise HTTPException(status_code=400, detail="Deploy model before setting it active")

    db.execute(text("UPDATE ml_models SET is_active = 0"))

    db.execute(
        text(
            """
            UPDATE ml_models
            SET is_active = 1,
                activated_at = :activated_at,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "activated_at": now_text(),
            "updated_at": now_text(),
            "id": model_id,
        },
    )

    db.commit()

    return {
        "active": True,
        "model": serialize_model(get_model_or_404(db, model_id)),
    }


@router.patch("/models/{model_id}/deactivate")
def deactivate_ml_model(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    get_model_or_404(db, model_id)

    db.execute(
        text(
            """
            UPDATE ml_models
            SET is_active = 0,
                updated_at = :updated_at
            WHERE id = :id
            """
        ),
        {
            "updated_at": now_text(),
            "id": model_id,
        },
    )

    db.commit()

    return {
        "deactivated": True,
        "model": serialize_model(get_model_or_404(db, model_id)),
    }


@router.get("/applications/ready")
def list_ml_ready_applications(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    applications = db.query(Application).order_by(Application.id.desc()).all()

    data = []

    for application in applications:
        features = application_to_features(application)

        latest_prediction = db.execute(
            text(
                """
                SELECT *
                FROM ml_predictions
                WHERE application_id = :application_id
                ORDER BY id DESC
                LIMIT 1
                """
            ),
            {"application_id": application.id},
        ).mappings().first()

        data.append(
            {
                "id": application.id,
                "application_id": application.id,
                "applicant_name": f"{safe_text(getattr(application, 'first_name', ''))} {safe_text(getattr(application, 'last_name', ''))}".strip() or "Unknown Applicant",
                "occupation": features["occupation"],
                "monthly_income": features["monthly_income"],
                "status": safe_text(getattr(application, "status", "draft"), "draft"),
                "ml_ready": bool(features["occupation"] and features["monthly_income"]),
                "latest_prediction": dict(latest_prediction) if latest_prediction else None,
            }
        )

    return data


@router.post("/applications/{application_id}/predict")
def predict_application_with_active_model(
    application_id: int,
    payload: dict = Body(default={}),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    application = get_application_or_404(db, application_id)
    active_model = get_active_model(db)

    if not active_model:
        raise HTTPException(status_code=400, detail="No active model found. Deploy and set active a model first.")

    features = application_to_features(application)

    # Optional payload can override fresh visible Apply page values later.
    for key in [
        "age",
        "monthly_income",
        "occupation",
        "phone",
        "email",
        "address",
    ]:
        if key in payload and payload[key] not in [None, ""]:
            features[key] = payload[key]

    features["has_occupation"] = 1 if safe_text(features.get("occupation")) else 0
    features["has_phone"] = 1 if safe_text(features.get("phone")) else 0
    features["has_email"] = 1 if safe_text(features.get("email")) else 0
    features["has_address"] = 1 if safe_text(features.get("address")) else 0

    prediction = try_uploaded_model_prediction(active_model.get("model_path"), features)

    if not prediction:
        prediction = baseline_prediction(features)
        prediction["reason"] = f"{prediction['reason']} Active uploaded model could not be used safely with current features, so baseline fallback logic was applied."

    created_at = now_text()

    result = db.execute(
        text(
            """
            INSERT INTO ml_predictions (
                application_id,
                model_id,
                result,
                risk_level,
                confidence,
                reason,
                features_json,
                created_at
            )
            VALUES (
                :application_id,
                :model_id,
                :result,
                :risk_level,
                :confidence,
                :reason,
                :features_json,
                :created_at
            )
            """
        ),
        {
            "application_id": application.id,
            "model_id": active_model["id"],
            "result": prediction["result"],
            "risk_level": prediction["risk_level"],
            "confidence": prediction["confidence"],
            "reason": prediction["reason"],
            "features_json": json.dumps(features, ensure_ascii=False),
            "created_at": created_at,
        },
    )

    db.commit()

    prediction_id = result.lastrowid

    return {
        "prediction_id": prediction_id,
        "application_id": application.id,
        "model": serialize_model(active_model),
        "result": prediction["result"],
        "risk_level": prediction["risk_level"],
        "confidence": prediction["confidence"],
        "reason": prediction["reason"],
        "features": features,
        "mapped_features": prediction.get("mapped_features", {}),
        "created_at": created_at,
    }




def resolve_mlops_file_path(file_path: str | None) -> Path | None:
    if not file_path:
        return None

    raw = Path(str(file_path).strip().replace("\\", "/"))

    candidates = [
        raw,
        Path.cwd() / raw,
        Path("uploads") / raw,
        Path.cwd() / "uploads" / raw,
        Path.cwd() / "uploads" / raw.name,
    ]

    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_file():
                return candidate
        except Exception:
            pass

    try:
        matches = list(Path("uploads").rglob(raw.name))
        matches = [item for item in matches if item.is_file()]
        if matches:
            matches.sort(key=lambda item: item.stat().st_mtime, reverse=True)
            return matches[0]
    except Exception:
        pass

    return None


def safe_string(value, max_length: int = 500) -> str:
    try:
        text_value = str(value)
    except Exception:
        text_value = f"<unreadable {type(value).__name__}>"

    text_value = text_value.replace("\n", " ").replace("\r", " ").strip()

    if len(text_value) > max_length:
        return text_value[:max_length] + "..."

    return text_value


def safe_list(value, max_items: int = 50) -> list[str]:
    try:
        if value is None:
            return []

        items = list(value)

        return [safe_string(item, 150) for item in items[:max_items]]
    except Exception:
        return []


def safe_dict_preview(value, max_items: int = 25) -> dict:
    try:
        if not isinstance(value, dict):
            return {}

        result = {}

        for index, (key, item) in enumerate(value.items()):
            if index >= max_items:
                break

            result[safe_string(key, 150)] = safe_string(item, 250)

        return result
    except Exception:
        return {}


def preview_csv_file(path: Path, limit: int = 10) -> dict:
    import csv

    rows = []

    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as file:
        reader = csv.DictReader(file)
        columns = reader.fieldnames or []

        for index, row in enumerate(reader):
            if index >= limit:
                break

            rows.append({column: row.get(column, "") for column in columns})

    return {
        "columns": columns,
        "rows": rows,
    }


def preview_dataset_file(path: Path, limit: int = 10) -> dict:
    suffix = path.suffix.lower()

    if suffix == ".csv":
        preview = preview_csv_file(path, limit)

        return {
            "preview_type": "table",
            "columns": preview["columns"],
            "rows": preview["rows"],
            "message": "Dataset preview loaded successfully.",
        }

    try:
        import pandas as pd

        if suffix in [".xlsx", ".xls"]:
            frame = pd.read_excel(path, nrows=limit)
        elif suffix == ".json":
            frame = pd.read_json(path).head(limit)
        elif suffix in [".txt", ".tsv"]:
            frame = pd.read_csv(path, sep=None, engine="python", nrows=limit)
        else:
            return {
                "preview_type": "unsupported",
                "columns": [],
                "rows": [],
                "message": "Dataset is stored, but this file type does not support table preview.",
            }

        frame = frame.fillna("")

        return {
            "preview_type": "table",
            "columns": [str(column) for column in frame.columns],
            "rows": frame.astype(str).to_dict(orient="records"),
            "message": "Dataset preview loaded successfully.",
        }
    except Exception as error:
        return {
            "preview_type": "error",
            "columns": [],
            "rows": [],
            "message": safe_string(error, 500),
        }


def load_model_object_safely(path: Path):
    errors = []

    try:
        import joblib
        return joblib.load(path), "joblib", errors
    except Exception as error:
        errors.append(f"joblib load failed: {safe_string(error, 300)}")

    try:
        with path.open("rb") as file:
            return pickle.load(file), "pickle", errors
    except Exception as error:
        errors.append(f"pickle load failed: {safe_string(error, 300)}")

    return None, "metadata_only", errors


def inspect_model_file(path: Path, limit: int = 10) -> dict:
    suffix = path.suffix.lower()
    file_size_mb = round(path.stat().st_size / (1024 * 1024), 4)

    model_object, loader, load_errors = load_model_object_safely(path)

    if model_object is None:
        return {
            "preview_type": "model_metadata_only",
            "loader": loader,
            "object_type": "Stored model file",
            "object_module": "unknown",
            "can_predict": False,
            "can_predict_proba": False,
            "classes": [],
            "n_features_in": None,
            "feature_names_in": [],
            "parameters": {},
            "table_preview": None,
            "dict_keys": [],
            "file_extension": suffix,
            "file_size_mb": file_size_mb,
            "load_status": "not_loaded",
            "compatibility_note": "Model file is stored successfully, but this backend could not fully load the model object. Install matching training dependencies or re-export the model.",
            "load_errors": load_errors,
            "message": "Model file metadata is available safely.",
        }

    object_to_inspect = model_object
    dict_keys = []

    if isinstance(model_object, dict):
        dict_keys = [safe_string(key, 150) for key in list(model_object.keys())[:50]]

        for key in ["model", "estimator", "classifier", "pipeline", "trained_model"]:
            possible_model = model_object.get(key)

            if possible_model is not None and callable(getattr(possible_model, "predict", None)):
                object_to_inspect = possible_model
                break

    object_type = safe_string(type(object_to_inspect).__name__, 150)
    object_module = safe_string(type(object_to_inspect).__module__, 200)

    try:
        can_predict = callable(getattr(object_to_inspect, "predict", None))
    except Exception:
        can_predict = False

    try:
        can_predict_proba = callable(getattr(object_to_inspect, "predict_proba", None))
    except Exception:
        can_predict_proba = False

    try:
        classes = safe_list(getattr(object_to_inspect, "classes_", None), 50)
    except Exception:
        classes = []

    try:
        feature_names = safe_list(getattr(object_to_inspect, "feature_names_in_", None), 100)
    except Exception:
        feature_names = []

    try:
        raw_n_features = getattr(object_to_inspect, "n_features_in_", None)
        n_features = int(raw_n_features) if raw_n_features is not None else None
    except Exception:
        n_features = None

    try:
        get_params = getattr(object_to_inspect, "get_params", None)
        parameters = safe_dict_preview(get_params(), 25) if callable(get_params) else {}
    except Exception:
        parameters = {}

    table_preview = None

    try:
        if hasattr(object_to_inspect, "head") and hasattr(object_to_inspect, "columns"):
            frame = object_to_inspect.head(limit).fillna("")
            table_preview = {
                "columns": [safe_string(column, 150) for column in frame.columns],
                "rows": frame.astype(str).to_dict(orient="records"),
            }
    except Exception:
        table_preview = None

    return {
        "preview_type": "model_object",
        "loader": loader,
        "object_type": object_type,
        "object_module": object_module,
        "can_predict": can_predict,
        "can_predict_proba": can_predict_proba,
        "classes": classes,
        "n_features_in": n_features,
        "feature_names_in": feature_names,
        "parameters": parameters,
        "table_preview": table_preview,
        "dict_keys": dict_keys,
        "file_extension": suffix,
        "file_size_mb": file_size_mb,
        "load_status": "loaded",
        "compatibility_note": "Model object loaded safely. If Apply page feature shape does not match this model, prediction will use fallback logic instead of crashing.",
        "load_errors": load_errors,
        "message": "Model object inspected successfully.",
    }



def mlops_view_safe_string(value, max_length: int = 500) -> str:
    try:
        text_value = str(value)
    except Exception:
        text_value = f"<unreadable {type(value).__name__}>"

    text_value = text_value.replace("\n", " ").replace("\r", " ").strip()

    if len(text_value) > max_length:
        return text_value[:max_length] + "..."

    return text_value


def mlops_view_safe_number(value, default: float = 0.0) -> float:
    try:
        if value is None:
            return default

        return float(value)
    except Exception:
        return default


def mlops_view_safe_list(value, max_items: int = 100) -> list[str]:
    try:
        if value is None:
            return []

        return [mlops_view_safe_string(item, 150) for item in list(value)[:max_items]]
    except Exception:
        return []


def mlops_view_safe_params(value, max_items: int = 30) -> dict:
    try:
        if not isinstance(value, dict):
            return {}

        result = {}

        for index, (key, item) in enumerate(value.items()):
            if index >= max_items:
                break

            result[mlops_view_safe_string(key, 150)] = mlops_view_safe_string(item, 250)

        return result
    except Exception:
        return {}


def mlops_view_load_model(path: Path):
    errors = []

    try:
        import joblib

        return joblib.load(path), "joblib", errors
    except Exception as error:
        errors.append(f"joblib load failed: {mlops_view_safe_string(error, 350)}")

    try:
        import pickle

        with path.open("rb") as file:
            return pickle.load(file), "pickle", errors
    except Exception as error:
        errors.append(f"pickle load failed: {mlops_view_safe_string(error, 350)}")

    return None, "metadata_only", errors


def mlops_view_predictable_object(model_object):
    if model_object is None:
        return None

    try:
        if callable(getattr(model_object, "predict", None)):
            return model_object
    except Exception:
        pass

    if isinstance(model_object, dict):
        for key in ["model", "estimator", "classifier", "pipeline", "trained_model"]:
            possible_model = model_object.get(key)

            try:
                if possible_model is not None and callable(getattr(possible_model, "predict", None)):
                    return possible_model
            except Exception:
                pass

    return model_object


def mlops_view_pipeline_steps(model_object) -> list[dict]:
    steps = []

    try:
        raw_steps = getattr(model_object, "steps", None)

        if raw_steps is None:
            return []

        for name, step in list(raw_steps):
            step_params = {}

            try:
                get_params = getattr(step, "get_params", None)
                step_params = mlops_view_safe_params(get_params(), 12) if callable(get_params) else {}
            except Exception:
                step_params = {}

            steps.append(
                {
                    "name": mlops_view_safe_string(name, 120),
                    "type": mlops_view_safe_string(type(step).__name__, 150),
                    "module": mlops_view_safe_string(type(step).__module__, 200),
                    "parameters": step_params,
                }
            )
    except Exception:
        return []

    return steps


def mlops_view_final_estimator(model_object):
    try:
        raw_steps = getattr(model_object, "steps", None)

        if raw_steps is not None and len(list(raw_steps)) > 0:
            return list(raw_steps)[-1][1]
    except Exception:
        pass

    return model_object


def mlops_view_feature_names(model_object, final_estimator) -> list[str]:
    for item in [model_object, final_estimator]:
        try:
            raw_features = getattr(item, "feature_names_in_", None)

            if raw_features is not None:
                features = [mlops_view_safe_string(feature, 150) for feature in list(raw_features)]
                if len(features) > 0:
                    return features
        except Exception:
            pass

    try:
        raw_n_features = getattr(model_object, "n_features_in_", None)

        if raw_n_features is None:
            raw_n_features = getattr(final_estimator, "n_features_in_", None)

        if raw_n_features is not None:
            return [f"feature_{index + 1}" for index in range(int(raw_n_features))]
    except Exception:
        pass

    return []


def mlops_view_classes(model_object, final_estimator) -> list[str]:
    for item in [model_object, final_estimator]:
        try:
            raw_classes = getattr(item, "classes_", None)

            if raw_classes is not None:
                classes = mlops_view_safe_list(raw_classes, 100)
                if len(classes) > 0:
                    return classes
        except Exception:
            pass

    return []


def mlops_view_feature_default(feature_name: str) -> float:
    name = str(feature_name or "").strip().lower()

    if name in ["age", "applicant_age"]:
        return 30

    if name in ["monthly_income", "income", "salary", "monthly_salary"]:
        return 25000

    if name in ["loan_amount", "requested_loan_amount", "amount"]:
        return 150000

    if name in ["credit_score", "cibil_score", "score"]:
        return 650

    if name in ["employment_years", "years_employed", "work_experience", "experience_years"]:
        return 2

    if name in ["existing_debt", "current_debt", "debt"]:
        return 0

    if name.startswith("has_"):
        return 1

    return 0


def mlops_view_coefficients(final_estimator, feature_names: list[str], classes: list[str]) -> list[dict]:
    try:
        import numpy as np

        raw_coef = getattr(final_estimator, "coef_", None)

        if raw_coef is None:
            return []

        coef_array = np.asarray(raw_coef)

        if coef_array.ndim == 1:
            coef_array = coef_array.reshape(1, -1)

        rows = []

        for class_index in range(coef_array.shape[0]):
            class_label = classes[class_index] if class_index < len(classes) else f"class_{class_index}"

            for feature_index in range(coef_array.shape[1]):
                feature_name = (
                    feature_names[feature_index]
                    if feature_index < len(feature_names)
                    else f"feature_{feature_index + 1}"
                )

                rows.append(
                    {
                        "class": mlops_view_safe_string(class_label, 100),
                        "feature": mlops_view_safe_string(feature_name, 150),
                        "coefficient": float(coef_array[class_index][feature_index]),
                    }
                )

        return rows
    except Exception:
        return []


def mlops_view_feature_importance(final_estimator, feature_names: list[str]) -> list[dict]:
    try:
        import numpy as np

        raw_importance = getattr(final_estimator, "feature_importances_", None)

        if raw_importance is None:
            return []

        importance_array = np.asarray(raw_importance).reshape(-1)

        rows = []

        for index in range(len(importance_array)):
            feature_name = feature_names[index] if index < len(feature_names) else f"feature_{index + 1}"

            rows.append(
                {
                    "feature": mlops_view_safe_string(feature_name, 150),
                    "importance": float(importance_array[index]),
                }
            )

        rows.sort(key=lambda item: item["importance"], reverse=True)

        return rows
    except Exception:
        return []


def mlops_view_sample_prediction(model_object, feature_names: list[str], classes: list[str]) -> dict:
    if not feature_names:
        return {
            "sample_input": {},
            "sample_prediction": None,
            "sample_probabilities": {},
        }

    sample_input = {
        feature_name: mlops_view_feature_default(feature_name)
        for feature_name in feature_names
    }

    try:
        try:
            import pandas as pd

            model_input = pd.DataFrame([sample_input], columns=feature_names)
        except Exception:
            model_input = [[sample_input[feature_name] for feature_name in feature_names]]

        raw_prediction = model_object.predict(model_input)[0]

        sample_probabilities = {}

        predict_proba = getattr(model_object, "predict_proba", None)

        if callable(predict_proba):
            try:
                raw_proba = list(predict_proba(model_input)[0])

                for index, value in enumerate(raw_proba):
                    label = classes[index] if index < len(classes) else f"class_{index}"
                    sample_probabilities[mlops_view_safe_string(label, 100)] = round(float(value) * 100, 4)
            except Exception:
                sample_probabilities = {}

        return {
            "sample_input": sample_input,
            "sample_prediction": mlops_view_safe_string(raw_prediction, 150),
            "sample_probabilities": sample_probabilities,
        }
    except Exception as error:
        return {
            "sample_input": sample_input,
            "sample_prediction": None,
            "sample_probabilities": {},
            "sample_error": mlops_view_safe_string(error, 350),
        }


def inspect_model_file_perfect(path: Path, limit: int = 10) -> dict:
    suffix = path.suffix.lower()
    file_size_mb = round(path.stat().st_size / (1024 * 1024), 4)

    model_object, loader, load_errors = mlops_view_load_model(path)

    raw_file_notice = (
        "This is a serialized binary model file, not a CSV table. "
        "SmartLoan AI is showing an interpreted MLOps inspection view of the file."
    )

    if model_object is None:
        return {
            "preview_type": "model_file_view",
            "view_mode": "metadata_only",
            "loader": loader,
            "object_type": "Stored binary model file",
            "object_module": "unknown",
            "can_predict": False,
            "can_predict_proba": False,
            "classes": [],
            "n_features_in": None,
            "feature_names_in": [],
            "required_features": [],
            "pipeline_steps": [],
            "final_estimator": {},
            "parameters": {},
            "coefficient_table": [],
            "feature_importance_table": [],
            "sample_input": {},
            "sample_prediction": None,
            "sample_probabilities": {},
            "table_preview": None,
            "dict_keys": [],
            "file_extension": suffix,
            "file_size_mb": file_size_mb,
            "load_status": "not_loaded",
            "raw_file_notice": raw_file_notice,
            "compatibility_note": "The file is stored successfully, but this backend environment could not fully load it. Install the same training dependencies/versions or re-export the model.",
            "load_errors": load_errors,
            "message": "Model file metadata is available safely.",
        }

    object_to_inspect = mlops_view_predictable_object(model_object)
    final_estimator = mlops_view_final_estimator(object_to_inspect)
    pipeline_steps = mlops_view_pipeline_steps(object_to_inspect)
    feature_names = mlops_view_feature_names(object_to_inspect, final_estimator)
    classes = mlops_view_classes(object_to_inspect, final_estimator)

    try:
        can_predict = callable(getattr(object_to_inspect, "predict", None))
    except Exception:
        can_predict = False

    try:
        can_predict_proba = callable(getattr(object_to_inspect, "predict_proba", None))
    except Exception:
        can_predict_proba = False

    n_features = len(feature_names) if len(feature_names) > 0 else None

    try:
        raw_n_features = getattr(object_to_inspect, "n_features_in_", None)

        if raw_n_features is not None:
            n_features = int(raw_n_features)
    except Exception:
        pass

    try:
        get_params = getattr(object_to_inspect, "get_params", None)
        parameters = mlops_view_safe_params(get_params(), 35) if callable(get_params) else {}
    except Exception:
        parameters = {}

    dict_keys = []

    if isinstance(model_object, dict):
        dict_keys = [mlops_view_safe_string(key, 150) for key in list(model_object.keys())[:50]]

    coefficient_table = mlops_view_coefficients(final_estimator, feature_names, classes)
    feature_importance_table = mlops_view_feature_importance(final_estimator, feature_names)
    sample_prediction = mlops_view_sample_prediction(object_to_inspect, feature_names, classes)

    table_preview = None

    try:
        if hasattr(model_object, "head") and hasattr(model_object, "columns"):
            frame = model_object.head(limit).fillna("")
            table_preview = {
                "columns": [mlops_view_safe_string(column, 150) for column in frame.columns],
                "rows": frame.astype(str).to_dict(orient="records"),
            }
    except Exception:
        table_preview = None

    return {
        "preview_type": "model_file_view",
        "view_mode": "interpreted_model_view",
        "loader": loader,
        "object_type": mlops_view_safe_string(type(object_to_inspect).__name__, 150),
        "object_module": mlops_view_safe_string(type(object_to_inspect).__module__, 200),
        "can_predict": can_predict,
        "can_predict_proba": can_predict_proba,
        "classes": classes,
        "n_features_in": n_features,
        "feature_names_in": feature_names,
        "required_features": feature_names,
        "pipeline_steps": pipeline_steps,
        "final_estimator": {
            "type": mlops_view_safe_string(type(final_estimator).__name__, 150),
            "module": mlops_view_safe_string(type(final_estimator).__module__, 200),
        },
        "parameters": parameters,
        "coefficient_table": coefficient_table,
        "feature_importance_table": feature_importance_table,
        "sample_input": sample_prediction.get("sample_input", {}),
        "sample_prediction": sample_prediction.get("sample_prediction"),
        "sample_probabilities": sample_prediction.get("sample_probabilities", {}),
        "sample_error": sample_prediction.get("sample_error"),
        "table_preview": table_preview,
        "dict_keys": dict_keys,
        "file_extension": suffix,
        "file_size_mb": file_size_mb,
        "load_status": "loaded",
        "raw_file_notice": raw_file_notice,
        "compatibility_note": "Model object loaded and converted into a readable MLOps inspection view.",
        "load_errors": load_errors,
        "message": "Model file inspected successfully.",
    }



@router.get("/models/{model_id}/dataset-preview")
def get_dataset_preview(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    model = get_model_or_404(db, model_id)

    file_path = resolve_mlops_file_path(model.get("dataset_path"))

    if not file_path:
        raise HTTPException(status_code=404, detail="Dataset file not found for this model.")

    preview = preview_dataset_file(file_path)

    return {
        "model_id": model_id,
        "file_name": model.get("dataset_original_name") or file_path.name,
        "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 4),
        **preview,
    }


@router.get("/models/{model_id}/model-preview")
def get_model_preview(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    model = get_model_or_404(db, model_id)

    file_path = resolve_mlops_file_path(model.get("model_path"))

    if not file_path:
        raise HTTPException(status_code=404, detail="Model file not found for this model.")

    preview = inspect_model_file_perfect(file_path)

    return {
        "model_id": model_id,
        "file_name": model.get("model_original_name") or file_path.name,
        "file_size_mb": round(file_path.stat().st_size / (1024 * 1024), 4),
        **preview,
    }



@router.get("/predictions")
def list_prediction_history(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    rows = db.execute(
        text(
            """
            SELECT p.*, m.model_name, m.version
            FROM ml_predictions p
            LEFT JOIN ml_models m ON m.id = p.model_id
            ORDER BY p.id DESC
            """
        )
    ).mappings().all()

    return [dict(row) for row in rows]




def file_size_bytes(path_value: str | None) -> int:
    if not path_value:
        return 0

    path = Path(path_value)

    if not path.exists():
        return 0

    try:
        return path.stat().st_size
    except Exception:
        return 0


def preview_csv_dataset(path_value: str) -> dict:
    path = Path(path_value)

    if not path.exists():
        return {
            "available": False,
            "message": "Dataset file is missing from storage.",
            "columns": [],
            "rows": [],
        }

    rows = []

    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as file:
            sample = file.read(4096)
            file.seek(0)

            try:
                dialect = csv.Sniffer().sniff(sample)
            except Exception:
                dialect = csv.excel

            reader = csv.DictReader(file, dialect=dialect)

            columns = list(reader.fieldnames or [])

            for row in itertools.islice(reader, 10):
                rows.append({column: row.get(column, "") for column in columns})

            return {
                "available": True,
                "file_type": "csv",
                "columns": columns,
                "rows": rows,
                "message": f"Showing first {len(rows)} rows from uploaded dataset.",
            }

    except UnicodeDecodeError:
        with open(path, "r", encoding="latin-1", newline="") as file:
            reader = csv.DictReader(file)
            columns = list(reader.fieldnames or [])

            for row in itertools.islice(reader, 10):
                rows.append({column: row.get(column, "") for column in columns})

            return {
                "available": True,
                "file_type": "csv",
                "columns": columns,
                "rows": rows,
                "message": f"Showing first {len(rows)} rows from uploaded dataset.",
            }

    except Exception as error:
        return {
            "available": False,
            "file_type": "csv",
            "columns": [],
            "rows": [],
            "message": f"CSV preview failed: {error}",
        }


def preview_json_dataset(path_value: str) -> dict:
    path = Path(path_value)

    if not path.exists():
        return {
            "available": False,
            "message": "Dataset file is missing from storage.",
            "columns": [],
            "rows": [],
        }

    try:
        data = json.loads(path.read_text(encoding="utf-8-sig", errors="ignore"))

        if isinstance(data, list):
            rows = data[:10]

            if rows and isinstance(rows[0], dict):
                columns = list(rows[0].keys())

                return {
                    "available": True,
                    "file_type": "json",
                    "columns": columns,
                    "rows": rows,
                    "message": f"Showing first {len(rows)} JSON rows.",
                }

            return {
                "available": True,
                "file_type": "json",
                "columns": ["value"],
                "rows": [{"value": str(item)} for item in rows],
                "message": f"Showing first {len(rows)} JSON values.",
            }

        if isinstance(data, dict):
            rows = [{"key": key, "value": str(value)} for key, value in list(data.items())[:10]]

            return {
                "available": True,
                "file_type": "json",
                "columns": ["key", "value"],
                "rows": rows,
                "message": "Showing first 10 JSON key/value pairs.",
            }

        return {
            "available": False,
            "file_type": "json",
            "columns": [],
            "rows": [],
            "message": "JSON file is not list/dict format.",
        }

    except Exception as error:
        return {
            "available": False,
            "file_type": "json",
            "columns": [],
            "rows": [],
            "message": f"JSON preview failed: {error}",
        }


def preview_excel_dataset(path_value: str) -> dict:
    path = Path(path_value)

    if not path.exists():
        return {
            "available": False,
            "message": "Dataset file is missing from storage.",
            "columns": [],
            "rows": [],
        }

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active

        values = list(itertools.islice(sheet.iter_rows(values_only=True), 11))

        if not values:
            return {
                "available": True,
                "file_type": "excel",
                "columns": [],
                "rows": [],
                "message": "Excel file is empty.",
            }

        columns = [str(item or f"Column {index + 1}") for index, item in enumerate(values[0])]
        rows = []

        for row_values in values[1:]:
            rows.append({
                columns[index]: "" if value is None else str(value)
                for index, value in enumerate(row_values)
                if index < len(columns)
            })

        return {
            "available": True,
            "file_type": "excel",
            "columns": columns,
            "rows": rows,
            "message": f"Showing first {len(rows)} rows from first Excel sheet.",
        }

    except Exception as error:
        return {
            "available": False,
            "file_type": "excel",
            "columns": [],
            "rows": [],
            "message": f"Excel preview failed: {error}",
        }


def preview_dataset_file(path_value: str | None, original_name: str | None) -> dict:
    if not path_value:
        return {
            "available": False,
            "file_name": original_name or "",
            "file_size_bytes": 0,
            "columns": [],
            "rows": [],
            "message": "No dataset uploaded for this model.",
        }

    suffix = Path(original_name or path_value).suffix.lower()

    if suffix == ".csv" or suffix == ".txt":
        preview = preview_csv_dataset(path_value)
    elif suffix == ".json":
        preview = preview_json_dataset(path_value)
    elif suffix in [".xlsx", ".xlsm", ".xls"]:
        preview = preview_excel_dataset(path_value)
    else:
        preview = {
            "available": False,
            "file_type": suffix.replace(".", "") or "unknown",
            "columns": [],
            "rows": [],
            "message": "Preview is supported for CSV, JSON, and Excel files.",
        }

    preview["file_name"] = original_name or Path(path_value).name
    preview["file_size_bytes"] = file_size_bytes(path_value)

    return preview


def load_model_artifact(path_value: str):
    path = Path(path_value)

    try:
        import joblib
        return joblib.load(path)
    except Exception:
        pass

    with open(path, "rb") as file:
        return pickle.load(file)


def preview_loaded_model_object(model_object: Any) -> dict:
    model_class = model_object.__class__.__name__
    model_module = model_object.__class__.__module__

    metadata = {
        "model_class": model_class,
        "model_module": model_module,
        "n_features_in": getattr(model_object, "n_features_in_", None),
        "feature_names": list(getattr(model_object, "feature_names_in_", []) or []),
        "classes": [str(item) for item in list(getattr(model_object, "classes_", []) or [])],
        "params": {},
        "table_preview": None,
    }

    if hasattr(model_object, "get_params"):
        try:
            params = model_object.get_params()
            metadata["params"] = {
                str(key): str(value)
                for key, value in list(params.items())[:15]
            }
        except Exception:
            metadata["params"] = {}

    if isinstance(model_object, dict):
        rows = [{"key": str(key), "value": str(value)} for key, value in list(model_object.items())[:10]]

        metadata["table_preview"] = {
            "columns": ["key", "value"],
            "rows": rows,
        }

    elif isinstance(model_object, list):
        rows = [{"value": str(item)} for item in model_object[:10]]

        metadata["table_preview"] = {
            "columns": ["value"],
            "rows": rows,
        }

    return metadata


def preview_model_file(path_value: str | None, original_name: str | None) -> dict:
    if not path_value:
        return {
            "available": False,
            "file_name": original_name or "",
            "file_size_bytes": 0,
            "message": "No model file uploaded.",
            "metadata": None,
        }

    path = Path(path_value)

    if not path.exists():
        return {
            "available": False,
            "file_name": original_name or path.name,
            "file_size_bytes": 0,
            "message": "Model file is missing from storage.",
            "metadata": None,
        }

    try:
        model_object = load_model_artifact(path_value)
        metadata = preview_loaded_model_object(model_object)

        return {
            "available": True,
            "file_name": original_name or path.name,
            "file_size_bytes": file_size_bytes(path_value),
            "message": "Model file loaded successfully. Showing metadata.",
            "metadata": metadata,
        }

    except Exception as error:
        return {
            "available": False,
            "file_name": original_name or path.name,
            "file_size_bytes": file_size_bytes(path_value),
            "message": f"Model file exists, but preview/load failed safely: {error}",
            "metadata": None,
        }


@router.get("/models/{model_id}/artifacts")
def get_ml_model_artifact_preview(
    model_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    model = get_model_or_404(db, model_id)

    dataset_preview = preview_dataset_file(
        model.get("dataset_path"),
        model.get("dataset_original_name"),
    )

    model_preview = preview_model_file(
        model.get("model_path"),
        model.get("model_original_name"),
    )

    return {
        "model": serialize_model(model),
        "dataset_preview": dataset_preview,
        "model_preview": model_preview,
    }


@router.get("/dashboard")
def mlops_dashboard(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    check_admin(current_user)
    ensure_mlops_tables(db)

    total_models = db.execute(text("SELECT COUNT(*) AS c FROM ml_models")).mappings().first()["c"]
    active_models = db.execute(text("SELECT COUNT(*) AS c FROM ml_models WHERE is_active = 1")).mappings().first()["c"]
    total_predictions = db.execute(text("SELECT COUNT(*) AS c FROM ml_predictions")).mappings().first()["c"]
    high_risk = db.execute(text("SELECT COUNT(*) AS c FROM ml_predictions WHERE risk_level = 'high'")).mappings().first()["c"]

    applications = db.query(Application).all()

    ready_applications = 0

    for application in applications:
        features = application_to_features(application)

        if features["occupation"] and features["monthly_income"]:
            ready_applications += 1

    risk_rows = db.execute(
        text(
            """
            SELECT risk_level, COUNT(*) AS count
            FROM ml_predictions
            GROUP BY risk_level
            """
        )
    ).mappings().all()

    risk_distribution = {
        "low": 0,
        "medium": 0,
        "high": 0,
    }

    for row in risk_rows:
        risk_distribution[row["risk_level"]] = row["count"]

    return {
        "total_models": total_models,
        "active_models": active_models,
        "ready_applications": ready_applications,
        "total_predictions": total_predictions,
        "high_risk": high_risk,
        "risk_distribution": risk_distribution,
        "active_model": serialize_model(get_active_model(db)) if get_active_model(db) else None,
    }
